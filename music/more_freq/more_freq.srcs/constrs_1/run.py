import openpyxl
r_wb = openpyxl.load_workbook('pin.xlsx')
r_sheetnames = r_wb.sheetnames
r_ws = r_wb.get_sheet_by_name(r_sheetnames[0])

file = open("stopwatch.xdc", mode='w')

for col in range(3, 11):
    if (r_ws.cell(row=5, column=col).value != None):
        file.write("set_property IOSTANDARD LVCMOS33 [get_ports " +
                   str(r_ws.cell(row=5, column=col).value) + "]\n")
        file.write("set_property PACKAGE_PIN " + 
                   str(r_ws.cell(row=6, column=col).value) +
                   " [get_ports " +
                   str(r_ws.cell(row=5, column=col).value) + "]\n")
        print("write:" + str(r_ws.cell(row=5, column=col).value))
for col in range(2, 6):
    if (r_ws.cell(row=12, column=col).value != None):
        file.write("set_property IOSTANDARD LVCMOS33 [get_ports " +
                   str(r_ws.cell(row=12, column=col).value) + "]\n")
        file.write("set_property PACKAGE_PIN " + 
                   str(r_ws.cell(row=13, column=col).value) +
                   " [get_ports " +
                   str(r_ws.cell(row=12, column=col).value) + "]\n")
        print("write:" + str(r_ws.cell(row=12, column=col).value))
for col in range(3, 19):
    if (r_ws.cell(row=28, column=col).value != None):
        file.write("set_property IOSTANDARD LVCMOS33 [get_ports " +
                   str(r_ws.cell(row=28, column=col).value) + "]\n")
        file.write("set_property PACKAGE_PIN " + 
                   str(r_ws.cell(row=29, column=col).value) +
                   " [get_ports " +
                   str(r_ws.cell(row=28, column=col).value) + "]\n")
        print("write:" + str(r_ws.cell(row=28, column=col).value))
for col in range(3, 19):
    if (r_ws.cell(row=17, column=col).value != None):
        file.write("set_property IOSTANDARD LVCMOS33 [get_ports " +
                   str(r_ws.cell(row=17, column=col).value) + "]\n")
        file.write("set_property PACKAGE_PIN " + 
                   str(r_ws.cell(row=18, column=col).value) +
                   " [get_ports " +
                   str(r_ws.cell(row=17, column=col).value) + "]\n")
        print("write:" + str(r_ws.cell(row=17, column=col).value))
if (r_ws.cell(row=17, column=21).value != None):
    file.write("set_property IOSTANDARD LVCMOS33 [get_ports " +
               str(r_ws.cell(row=17, column=21).value) + "]\n")
    file.write("set_property PACKAGE_PIN " +
               str(r_ws.cell(row=18, column=21).value) +
               " [get_ports " +
               str(r_ws.cell(row=17, column=21).value) + "]\n")
    print("write:" + str(r_ws.cell(row=17, column=21).value))
# push button
if (r_ws.cell(row=7, column=16).value != None):
    file.write("set_property IOSTANDARD LVCMOS33 [get_ports " +
               str(r_ws.cell(row=7, column=16).value) + "]\n")
    file.write("set_property PACKAGE_PIN " +
               str(r_ws.cell(row=6, column=16).value) +
               " [get_ports " +
               str(r_ws.cell(row=7, column=16).value) + "]\n")
    print("write:" + str(r_ws.cell(row=7, column=16).value))
for col in range(15, 18):
    if (r_ws.cell(row=9, column=col).value != None):
        file.write("set_property IOSTANDARD LVCMOS33 [get_ports " +
                   str(r_ws.cell(row=9, column=col).value) + "]\n")
        file.write("set_property PACKAGE_PIN " + 
                   str(r_ws.cell(row=8, column=col).value) +
                   " [get_ports " +
                   str(r_ws.cell(row=9, column=col).value) + "]\n")
        print("write:" + str(r_ws.cell(row=9, column=col).value))
if (r_ws.cell(row=11, column=16).value != None):
    file.write("set_property IOSTANDARD LVCMOS33 [get_ports " +
               str(r_ws.cell(row=11, column=16).value) + "]\n")
    file.write("set_property PACKAGE_PIN " +
               str(r_ws.cell(row=10, column=16).value) +
               " [get_ports " +
               str(r_ws.cell(row=11, column=16).value) + "]\n")
    print("write:" + str(r_ws.cell(row=11, column=16).value))
file.close()
